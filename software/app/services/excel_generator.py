"""Workbook executivo que preserva premissas e trilha de cálculo."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.nfe_models import AuditSummary

SLATE = "0F172A"
SLATE_2 = "1E293B"
EMERALD = "22C55E"
WHITE = "F8FAFC"
MUTED = "CBD5E1"


def _title(ws, text: str, end_column: int = 6) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, text)
    cell.font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=SLATE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34


def _header(row) -> None:
    for cell in row:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=SLATE_2)
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _fit(ws, widths: dict[int, int] | None = None) -> None:
    widths = widths or {}
    for column in range(1, ws.max_column + 1):
        maximum = max((len(str(ws.cell(row, column).value or "")) for row in range(1, ws.max_row + 1)), default=8)
        ws.column_dimensions[get_column_letter(column)].width = widths.get(column, min(max(maximum + 2, 12), 44))
    ws.sheet_view.showGridLines = False


def generate_excel(summary: AuditSummary) -> bytes:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Resumo Executivo"
    _title(summary_ws, "Audita | Memória de auditoria", 4)
    summary_ws.append([])
    summary_ws.append(["Indicador", "Valor", "Premissa", "Status"])
    _header(summary_ws[3])
    rows = [
        ("Notas analisadas", summary.invoice_count, "XMLs de venda recebidos", "Informativo"),
        ("Receita auditada", float(summary.audited_revenue), "CFOP de saída", "Informativo"),
        ("Receita monofásica confirmada", float(summary.confirmed_monophase_revenue), "Catálogo MVP versionado", "Revisar com contador"),
        ("Potencial estimado", float(summary.estimated_overpayment), "Somente receita não segregada", "Não garantido"),
    ]
    for row in rows:
        summary_ws.append(row)
    for row in range(5, 8):
        summary_ws.cell(row, 2).number_format = 'R$ #,##0.00'
    summary_ws["A9"] = "Aviso"
    summary_ws["A9"].font = Font(bold=True, color="991B1B")
    summary_ws["B9"] = "Apoio técnico. Não é petição, parecer jurídico ou garantia de crédito. Validação do contador é obrigatória."
    summary_ws.merge_cells("B9:D9")
    summary_ws["B9"].alignment = Alignment(wrap_text=True)
    _fit(summary_ws, {1: 34, 2: 24, 3: 34, 4: 24})

    items = wb.create_sheet("Itens Auditados")
    _title(items, "Itens e trilha de evidências", 14)
    headers = [
        "Nota", "Item", "Produto", "NCM", "CFOP", "Valor", "CST PIS", "CST Cofins", "CSOSN (ICMS)",
        "Status", "Segmento", "Potencial estimado", "Regra/Fonte", "Pendências",
    ]
    items.append(headers)
    _header(items[2])
    for finding in summary.findings:
        items.append([
            finding.invoice_number,
            finding.item_number,
            finding.description,
            finding.ncm,
            finding.cfop,
            float(finding.product_value),
            finding.pis_cst,
            finding.cofins_cst,
            finding.csosn,
            finding.status.value,
            finding.segment,
            float(finding.estimated_overpayment),
            finding.evidence.legal_source if finding.evidence else "Sem regra no catálogo MVP",
            " | ".join(finding.pending_checks),
        ])
    total_row = items.max_row + 1
    items.cell(total_row, 11, "Total")
    items.cell(total_row, 12, f"=SUM(L3:L{total_row - 1})")
    items.cell(total_row, 11).font = Font(bold=True)
    items.cell(total_row, 12).font = Font(bold=True, color=EMERALD)
    for row in range(3, total_row + 1):
        items.cell(row, 6).number_format = 'R$ #,##0.00'
        items.cell(row, 12).number_format = 'R$ #,##0.00'
        for cell in items[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    items.auto_filter.ref = f"A2:N{total_row - 1}"
    items.freeze_panes = "A3"
    _fit(items, {3: 34, 13: 38, 14: 52})

    memory = wb.create_sheet("Memória PGDAS-D")
    _title(memory, "Premissas do cenário e cálculo", 5)
    memory.append([])
    memory.append(["Campo", "Valor", "Fórmula/Origem", "Validação", "Observação"])
    _header(memory[3])
    memory_rows = [
        ("RBT12", float(summary.context.rbt12), "Entrada do usuário", "Contador", "Anexo I"),
        ("Faixa", summary.rate.band, "Tabela do Anexo I", "Contador", "Faixas 1 a 5 no MVP"),
        ("Alíquota nominal", float(summary.rate.nominal_rate), "LC 123/2006", "Contador", ""),
        ("Parcela a deduzir", float(summary.rate.deduction), "LC 123/2006", "Contador", ""),
        ("Alíquota efetiva DAS", float(summary.rate.effective_das_rate), "(RBT12 × nominal − PD) ÷ RBT12", "Automático", ""),
        ("Fração PIS/Cofins", float(summary.rate.pis_cofins_share), "12,74% + 2,76% do DAS", "Contador", "Faixas suportadas"),
        ("Alíquota efetiva PIS/Cofins", float(summary.rate.effective_pis_cofins_rate), "Alíquota DAS × 15,5%", "Automático", ""),
        ("Receita já segregada?", "Sim" if summary.context.pgdas_segregated else "Não", "Entrada do usuário", "PGDAS-D", "Se sim, potencial zero"),
        ("Potencial estimado", float(summary.estimated_overpayment), "Soma da coluna Itens Auditados!L", "Contador", "Não garantido"),
    ]
    for row in memory_rows:
        memory.append(row)
    for row in (4, 7, 12):
        memory.cell(row, 2).number_format = 'R$ #,##0.00'
    for row in (6, 8, 9, 10):
        memory.cell(row, 2).number_format = "0.0000%"
    memory.freeze_panes = "A4"
    _fit(memory, {1: 34, 2: 24, 3: 42, 4: 20, 5: 36})

    thin = Side(style="thin", color="334155")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and cell.row > 1:
                    cell.border = Border(bottom=thin)
        ws.freeze_panes = ws.freeze_panes or "A3"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
