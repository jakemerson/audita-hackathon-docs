from io import BytesIO

from openpyxl import load_workbook
from pypdf import PdfReader

from app.core.nfe_parser import audit_xml_batch
from app.services.copilot_service import ask_copilot
from app.services.excel_generator import generate_excel
from app.services.openai_auditor import generate_audit_opinion, usable_key
from app.services.pdf_generator import generate_pdf
from tests.conftest import make_xml


def _summary(context):
    return audit_xml_batch([("fixture.xml", make_xml())], context)


def test_placeholder_key_is_never_usable():
    assert usable_key(None) is False
    assert usable_key("sua_chave_openai_aqui") is False


def test_ai_auditor_has_complete_local_fallback(context):
    opinion = generate_audit_opinion(_summary(context), "sua_chave_openai_aqui")
    assert opinion.source == "local"
    assert opinion.fundamentacao
    assert opinion.plano_acao
    assert "não é crédito garantido" in opinion.resumo


def test_copilot_rejects_guarantee_language():
    reply = ask_copilot("Você garante o Pix e o prazo?", "sua_chave_openai_aqui")
    assert reply.source == "local"
    assert "Não é possível garantir" in reply.answer
    assert reply.sources


def test_copilot_explains_auto_part_scope():
    reply = ask_copilot("Pastilha NCM 8708 entra?", None)
    assert "Lei 10.485/2002" in reply.answer
    assert "PGDAS-D" in reply.answer


def test_excel_reopens_with_three_auditable_sheets(context):
    payload = generate_excel(_summary(context))
    workbook = load_workbook(BytesIO(payload), data_only=False)
    assert workbook.sheetnames == ["Resumo Executivo", "Itens Auditados", "Memória PGDAS-D"]
    items = workbook["Itens Auditados"]
    assert items.cell(items.max_row, 12).value.startswith("=SUM(")
    assert "não é petição" in workbook["Resumo Executivo"]["B9"].value.lower()


def test_pdf_has_signature_and_extractable_safeguards(context):
    payload = generate_pdf(_summary(context))
    assert payload.startswith(b"%PDF")
    text = " ".join(page.extract_text() or "" for page in PdfReader(BytesIO(payload)).pages)
    assert "SIMULAÇÃO SINTÉTICA" in text
    assert "não é petição" in text
    assert "Contador responsável" in text
